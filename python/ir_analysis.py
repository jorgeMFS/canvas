#!python
#!/usr/bin/env python

import sys
import pandas as pd
import seaborn as sns
import networkx as nx
import math
import random
import statistics
import numpy as np
import openpyxl
from matplotlib.pyplot import colorbar
from matplotlib.patches import Ellipse
import matplotlib.pyplot as plt
import matplotlib as mpl
import os
from collections import Counter
from group_by_family import eprint, join_nc_taxa, read_file,match_files, read_genomic_type_path,read_organism_path, merge_measures,merge_measures2, final_merge,assert_merge, save_to_file

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


# Functions
def tree_average(nc_value_list):
    build_avg=[]
    elements_list=[]
    nc_list=[]
    ul=[]
    for x in nc_value_list:
        ul.append(x[1:])
    ul=list(set([item for sublist in ul for item in sublist])) 
    for elem in ul:
        avg_list=[]
        for x in nc_value_list:
            if elem in x:
                avg_list.append(x[0])
        build_avg.append([elem, sum(avg_list) / len(avg_list)])
    
    return build_avg

def floor_if_one(value):
    return 1 if value > 1 else value
 
def nc_process(path):
    list_nc = read_file(path)
    return [l[0:2]+[floor_if_one(float(min(l[2:])))] + [l[2:].index(min(l[2:]))] for l in list_nc]

def nc_process_no_floor(path):
    list_nc = read_file(path)
    return [l[0:2]+[float(min(l[2:]))] + [l[2:].index(min(l[2:]))] for l in list_nc]

def nc_diff(nc1_list,nc2_list): 
    diff_nc=[]
    for nc1, nc2 in zip(nc1_list,nc2_list):
        if nc1[1]==nc2[1]:
            diff_nc.append( nc1[0:2]+[float(nc1[2])-float(nc2[2])] + [nc1[3]] )
        else:
            eprint(nc1,nc2)
    return diff_nc

def process_meta_information(meta_information_list):
    metainfo=[]
    for meta_information in meta_information_list :
        strand = meta_information[2:4]
        strand.reverse()
        strand = [''.join(strand).replace("double","ds").replace("single","ss")]
        taxic_info = meta_information[0:2] + strand + meta_information[4:]
        metainfo.append(taxic_info)
    return metainfo

def process_organism_lists(organism_list_lists):
    unique_organims = [list(x) for x in set(tuple(x) for x in organism_list_lists)]
    unique_organims = [list(filter(None, lst)) for lst in unique_organims]
    unique_tax_list = list(set(flatten(unique_organims)))
    unique_taxa_occurance = list()
    for x in unique_tax_list:
        counter=0
        indices = set()
        for l in unique_organims:
            for i in range(0,len(l)):
                if x == l[i]:
                    counter+=1
                    indices.add(i)
                    
        taxa=[x, counter] + list(indices)
        unique_taxa_occurance.append(taxa)
        if len(list(indices))>1:
            eprint("Error With taxonomy")
    unique_taxa_occurance.sort(key=lambda x: x[1],reverse=True)

    return unique_taxa_occurance

def create_taxonomy_trees(organism_names, unique_list):
    top = [row for row in unique_list if row[1]>9]
    eprint(len(top))
    eprint(top[0])
    for row in top:
        index=row[2]
        tree_structure=[]
        for organism_name in organism_names:
            if row[0] in organism_name:
              tree_structure.append(organism_name)
        
        sucess = False
        end=max(len(elem) for elem in tree_structure)
        while sucess==False:
            sucess=create_taxonomy_tree([x[index:end] for x in tree_structure],row[0],row[0],"png")
            end-=1
        
def flatten(regular_list):
    return [item for sublist in regular_list for item in sublist]

def create_taxonomy_tree(meta_information_list, root, name, extension):
    save_string="../trees/"+name+"."+extension

    tree = dict()
    for virus_info in meta_information_list:
        for index in range(0,len(virus_info)):
            virus_info[index] = virus_info[index].replace(";", "").replace(":", "")
            if virus_info[index] not in tree.keys() and index +1 < len(virus_info):
                # if index +1 < len(virus_info):
                tree[virus_info[index]] = {virus_info[index + 1]:1}
                if index < len(virus_info) and index>0:
                    d1 = {virus_info[index]: 1}
                    tree[virus_info[index-1]].update(d1)
            else:
                if index!=0: #and virus_info[index] not in flatten(list(tree.values())) and virus_info[index] in tree.keys():
                    d1 = {virus_info[index]: 1}
                    tree[virus_info[index-1]].update(d1)

    newick_tree, const_bool = newickify(tree, root_node=root)
    if const_bool:
        t = Tree(newick_tree, quoted_node_names=True, format=1)
        ts, t = set_default_TreeStyle(t, True)
        t.set_style(ts)
        t.render(save_string,tree_style=ts, dpi=1000,h=120000, w=120000,units="px")
    return const_bool


def set_default_TreeStyle(tree, draw_nodes):
    ts = TreeStyle()
    ts.mode = "c"
    ts.arc_start = -180
    ts.arc_span = 180
    ts.root_opening_factor = 1
    ts.show_branch_length = False
    ts.show_branch_support = True
    ts.force_topology = False
    ts.show_leaf_name = False
    ts.min_leaf_separation=10
    ts.root_opening_factor=1
    ts.complete_branch_lines_when_necessary=True


    return ts, tree

def newickify(node_to_children, root_node) -> str:
    visited_nodes = set()

    def newick_render_node(name, distance: float) -> str:
        if name in visited_nodes:
            print(name)
        assert name not in visited_nodes, "Error: The tree may not be circular!"

        if name not in node_to_children:
            # Leafs
            return F'{name}:{distance}'
        else:
            # Nodes
            visited_nodes.add(name)
            children = node_to_children[name]
            children_strings = [newick_render_node(child, children[child]) for child in children.keys()]
            children_strings = ",".join(children_strings)
            return F'({children_strings}){name}:{distance}'

    newick_string = newick_render_node(root_node, 0) + ';'

    # Ensure no entries in the dictionary are left unused.
    assert visited_nodes == set(node_to_children.keys()), "Error: some nodes aren't in the tree"

    if visited_nodes != set(node_to_children.keys()):
        print("Error: some nodes aren't in the tree")
        return newick_string, False
    else:
        return newick_string, True

def bfs_edge_lst(graph, n):
    return list(nx.bfs_edges(graph, n))

def tree_from_edge_lst(elst):
    tree = {"Viruses": {}}
    for src, dst in elst:
        subt = recursive_search(tree, src)
        if subt:
            print(subt)
            subt[dst] = {}
    return tree

def recursive_search(dicts, key):
    if key in dicts:
        return dicts[key]
    for k, v in dicts.items():
        item = recursive_search(v, key)
        if item is not None:
            return item

def tree_to_newick(tree):
    items = []
    for k in tree.keys():
        s = ''
        if len(tree[k].keys()) > 0:
            subt = tree_to_newick(tree[k])
            if subt != '':
                s += '(' + str(subt) + ')'
        s += str(k)
        items.append(s)
    return ','.join(items)

def rgb(minimum, maximum, value):
    minimum, maximum = float(minimum), float(maximum)
    ratio = 2 * (value-minimum) / (maximum - minimum)
    b = int(max(0, 255*(1 - ratio)))
    r = int(max(0, 255*(ratio - 1)))
    g = 255 - b - r
    return r, g, b

def rgb2(minimum, maximum, value):
    value= (1-0)/(maximum-minimum)*(value-maximum)+1
    # color1=[255,255,0]
    color1=[0,0,255] # RGB for our 1st color (blue in this case).
    color2=[255,0,0]  #RGB for our 2nd color (red in this case)

    red   = int((color2[0] - color1[0]) * value + color1[0])     # Evaluated as -255*value + 255.
    green = int((color2[1] - color1[1]) * value + color1[1])      # Evaluates as 0.
    blue  = int((color2[2] - color1[2]) * value + color1[2])      # Evaluates as 255*value + 0.
    
    return red,green,blue


def colorbar(mn, mx, xvar, savepath):
    fig, ax = plt.subplots(figsize=(6, 1))
    fig.subplots_adjust(bottom=0.5)

    cnorm = mpl.colors.Normalize(vmin=mn, vmax=mx)
    cm1 = mpl.colors.LinearSegmentedColormap.from_list("MyCmapName",["b","r"])
    cb1 = mpl.colorbar.ColorbarBase(ax, cmap=cm1,
                                    norm=cnorm,
                                    orientation='horizontal')
    cb1.set_label(xvar)
    plt.savefig(savepath,  bbox_inches='tight',dpi=1000)
    plt.clf()
    plt.close(1)

def create_original_tree(meta_information_list, avg_list, root, name,sci,save_labels):

    save_string=name+".pdf"
    save_string_colorbar=name+"_colorbar.pdf"
    final_save=name+"_final"+".pdf"
    if sci:
        nc_val=[x[1] for x in avg_list]
    else:
        nc_val=[round(x[1],3) for x in avg_list]
    if not avg_list:
        avg_list=[[root,1.0],["V",0.0]]
    
    mx=max([x[1] for x in avg_list])
    mn=min([x[1] for x in avg_list])

    if mn == mx:
        mn=mn-0.001
        mx=mx+0.001
    if sci:
        mx=math.log(mx)
        if mn==0:
            mn=0.000001
        mn=math.log(mn)
    
    colorbar(mn, mx, save_labels, save_string_colorbar)
    leaf_val=[x[0] for x in avg_list]
    tree = dict()
    for virus_info in meta_information_list:
        for index in range(0,len(virus_info)):
            virus_info[index] = virus_info[index].replace(";", "").replace(":", "")
            if virus_info[index] not in tree.keys() and index +1 < len(virus_info):
                tree[virus_info[index]] = {virus_info[index + 1]:1}
                if index < len(virus_info) and index>0:
                    d1 = {virus_info[index]: 1}
                    tree[virus_info[index-1]].update(d1)
            else:
                if index!=0:
                    d1 = {virus_info[index]: 1}
                    tree[virus_info[index-1]].update(d1)

    newick_tree, const_bool = newickify(tree, root_node=root)

    if const_bool:
        t = Tree(newick_tree, quoted_node_names=True, format=1)
        ts, t = set_default_TreeStyle(t, False)
        t.set_style(ts)
        
        count=0
        for name,value in zip(leaf_val,nc_val):          
            matching_nodes = t.search_nodes(name=name)         
            if matching_nodes:
                dst=t.get_distance(root,matching_nodes[0])
                if sci:
                    if value==0:
                        rgb_color=rgb2(mn, mx, math.log(0.000001))
                        rgb_color=('#%02x%02x%02x' % rgb_color)
                        complexity = TextFace(value,fgcolor=rgb_color, fsize=200,bold=True)
                        change_tree_branch(matching_nodes, rgb_color, dst)
                    else:
                        rgb_color=rgb2(mn, mx, math.log(value))
                        rgb_color=('#%02x%02x%02x' % rgb_color)
                        complexity = TextFace("{:.2e}".format(value),fgcolor=rgb_color, fsize=200,bold=True)
                        change_tree_branch(matching_nodes, rgb_color, dst)
                else:
                    rgb_color=rgb2(mn, mx, value)
                    rgb_color=('#%02x%02x%02x' % rgb_color)
                    complexity = TextFace(value,fgcolor=rgb_color, fsize=200,bold=True)
                    change_tree_branch(matching_nodes, rgb_color, dst)
                    
                virus_name = TextFace(matching_nodes[0].name, fgcolor=rgb_color, fsize=200,bold=True)
                # matching_nodes[0].add_face(face=complexity, column=1, position="branch-bottom")
                matching_nodes[0].add_face(face=virus_name, column=1, position="branch-top")
            else:
                eprint("ERROR->", avg_list[count])
            count+=1
        t.render(save_string,tree_style=ts,  dpi=1000,h=120000, w=120000,units="px")

    return const_bool

def various_trees(meta_information_list,avg_list, root, name,sci,save_labels,save_folder):
    if not os.path.exists(save_folder):
        eprint("creating folder...")
        os.makedirs(save_folder)

    tax_g=["Super-Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus"]
    for x in range(1,len(meta_information_list[0])+1):
        met=[m[0:x] for m in meta_information_list]
        met=[m for m in met if "incertae_sedis" not in m[-1]]
        sv_name= save_folder+"/"+name+"_"+tax_g[x-1]
        avg_list_used = filter_based_on_list(met,avg_list)
        save_labels=name.split("_")[0]
        create_original_tree(met,avg_list_used, root,sv_name,sci,save_labels)

def various_trees_2(organism_names, average_join_var,save_folder):
    if not os.path.exists(save_folder):
        eprint("creating folder...")
        os.makedirs(save_folder)

    tax_g=["Super-Realm","Realm","Kingdom","Phylum","Class","Order","Family"]
    for x in range(0,len(tax_g)):
        save_folder2=save_folder+"/"+tax_g[x]
        if not os.path.exists(save_folder2):
            eprint("creating folder for trees in",tax_g[x],"...")
            os.makedirs(save_folder2)
        unique_list=list(set([m[x] for m in organism_names ]))
        pre_tree_list= [m[x:] for m in organism_names]

        for root in unique_list:
            if "incertae_sedis" not in root:
                savefolder=save_folder2+'/'+root
                tree_lst=[lst for lst in pre_tree_list if lst[0]==root]
                avg_list_used = filter_based_on_list(tree_lst,average_join_var)
                create_original_tree(tree_lst, avg_list_used, root, savefolder, True, "NC")
                
def split_avg_lst(avg_list_used):
    ir_0=[]
    ir_1=[]
    ir_2 =[]
    for lst in avg_list_used:
        ir_0.append([lst[0],lst[1]])
        ir_1.append([lst[0],lst[2]])
        ir_2.append([lst[0],lst[3]])
    return ir_0, ir_1, ir_2

def filter_based_on_list(meta_info, avg_list):
    unique_list=list(set(flatten(meta_info)))
    return [x for x in avg_list if x[0] in unique_list]

def change_tree_branch(node_list, rgb_color, dst):
    val=8-int(dst)
    style1=NodeStyle()
    style1["vt_line_width"] = val*40
    style1["hz_line_width"] = val*40
    style1["vt_line_color"] = rgb_color
    style1["hz_line_color"] = rgb_color
    for node in node_list:
        node.set_style(style1)


def attribute_color(node,color):
    node_style=NodeStyle()
    node_style["fgcolor"] = color
    textcolor = "Black"
    node_style["bgcolor"] = color
    if color=="Black":
        textcolor ="White"
    N =  TextFace(node.name, fgcolor=textcolor, fsize=50,bold=True)
    node.add_face(face=N, column=0)
    node.allow_face_overlap=True
    node.set_style(node_style)
    return node

def filter_from_ll(list_ir, ign_taxa):
    for taxa in ign_taxa:
        list_ir = [[ele for ele in sub if ele != taxa] for sub in list_ir]
    return list_ir

def list_fusion(list1,list2,list3):
    fused_list=[]
    for a,b,c in zip(list1,list2,list3):
        a_delta=a[1:]
        b_delta=b[1:]
        c_delta=c[1:]
        if a_delta==b_delta==c_delta:
            fused_list.append([float(a[0]),int(b[0]),float(c[0])] + a[1:])
        else:
            print("ERROR, some problem with merging! function 515")
            sys.exit()
    
    return fused_list
    


def create_features_list(fused_list, relative_path, save_name):
    if not os.path.exists(relative_path):
        eprint("creating folder for xls files...")
        os.makedirs(relative_path)
    
    header=["Normalized Compression","Sequence Length","GC-Content","Super-Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus", "Genome"]
    wb = openpyxl.Workbook()
    save_name = relative_path + "/" + save_name + ".xlsx"
    
    wb.save(filename = save_name)
    max_lenght = len(max(fused_list, key=len))
    save_list=[]
    for ind in range(3,max_lenght):
        unique_list = list(set([x[ind] for x in fused_list if len(x)>ind]))   
        unique_list = [ul for ul in unique_list if "incertae_sedis" not in ul and "unclassified" not in ul ]
        level_list=[]
        
        for un_ele in unique_list:
            l_values_nc=[]
            l_values_sl=[]
            l_values_gc=[]
            l_taxa=[]
            for l in fused_list:
                if len(l)>=ind and un_ele in l:
                    l_values_nc.append(l[0])
                    l_values_sl.append(l[1])
                    l_values_gc.append(l[2])
                    l_taxa=l[ind]
            avg_nc=sum(l_values_nc) / len(l_values_nc)
            avg_sl=sum(l_values_sl) / len(l_values_sl)
            avg_gc=sum(l_values_gc) / len(l_values_gc)
            level_list.append( [header[ind]]+[l_taxa]+[avg_nc]+[avg_sl]+[avg_gc])
        
        level_list.sort(key = lambda x: x[2], reverse=True) 
        save_list = save_list + level_list[:10]

    sheet_name="Feature Values"
    save_as_csv(save_name, sheet_name, save_list)
    wb = openpyxl.load_workbook(save_name)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(save_name)



def create_lists(list_ir, relative_path ,save_name):
    if not os.path.exists(relative_path):
        eprint("creating folder for xls files...")
        os.makedirs(relative_path)

    header=["None","Super-Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus","Species"]
    wb = openpyxl.Workbook()
    save_name = relative_path + "/" + save_name + ".xlsx"
    
    wb.save(filename = save_name)

    max_lenght = len(max(list_ir, key=len))
    for ind in range(2,max_lenght):
        unique_list = list(set([x[ind] for x in list_ir if len(x)>ind]))   
        unique_list=[ul for ul in unique_list if "incertae_sedis" not in ul and "unclassified" not in ul ]
        level_list=[]
        for un_ele in unique_list:
            l_values=[]
            l_taxa=[]
            for l in list_ir:
                if len(l)>=ind and un_ele in l:
                    l_values.append(l[0])
                    l_taxa=l[1:ind+1] 
            level_list.append([sum(l_values) / len(l_values)]+ l_taxa)
        
        level_list.sort(key = lambda x: x[0], reverse=True) 
        level_list.insert(0,header[0:ind+1])
        sheet_name=header[ind]
        save_as_csv(save_name, sheet_name, level_list)
    wb = openpyxl.load_workbook(save_name)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(save_name)
    print(f"{bcolors.OKBLUE}Top list successfully stored in:{bcolors.ENDC}",save_name)

def read_and_correct(file_path, organism_list):
    corrections=[line.split() for line in open(file_path, 'r')]
    item_list=[x[-1] for x in corrections]
    ind=0
    for item in item_list:
        for org in organism_list:
            if item in org:
                pos=org.index(item)
                if org[pos-1]==corrections[ind][0]:
                    org[pos:pos] = corrections[ind][1:-1]
        ind+=1
    return organism_list

def read_to_list(ignore_taxa_file):
    corrections=[line.replace("\n","") for line in open(ignore_taxa_file, 'r')]
    return corrections

def save_as_csv(path,sheetname, ls):  
    df = pd.DataFrame(ls)
    book = openpyxl.load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl",mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname)
    writer.save()

def filter_organism_list(organism_names):
    new_org=[]
    for org_des in organism_names:
        if len(org_des)>=8:
            not_detected=True
            for org in org_des:
                if "unclassified" in org:
                    not_detected=False
            if not_detected:
                new_org.append(org_des)
    new_org=[x[:-1] if len(x)>8  else x for x in new_org ] 
    return new_org

def filter_meta_info(meta_info):
    new_org=[]
    for meta in meta_info:
        if len(meta)>=12:
            not_detected=True
            for org in meta:
                if "unclassified" in org:
                    not_detected=False
            if not_detected:
                new_org.append(meta)
    l=[]
    for x in new_org:
        if len(x)>12:
            x.pop(len(x)-2)
        l.append(x)
    return l

def detect_wrong_pattern(list_of_taxa):
    taxa_list=["Viruses","viria","virae","viricota","viricetes","virales","viridae","virus","virus"]
    remove_list=[]
    for i in range(0,len(taxa_list)):
        tax_group=[l[i] for l in list_of_taxa if len(l)>i]
        [remove_list.append(tx) for tx in tax_group if "viricetidae" in tx]
    for x in list(set(remove_list)):
        print(x)
    sys.exit()
    
def join_variables(len_taxa_process, gc_taxa_process, nc_taxa_process,nc_taxa_ir0_process,nc_taxa_ir1_process, nc_taxa_ir2_process, diff_nc_taxa_process):
    join_var=[]
    for len_sq, gc_sq, nc_sq, ir0_sq, ir1_sq, ir2_sq, diff_sq  in zip(len_taxa_process, gc_taxa_process, nc_taxa_process,nc_taxa_ir0_process,nc_taxa_ir1_process, nc_taxa_ir2_process, diff_nc_taxa_process):
        if(len_sq[-1]==gc_sq[-1]==nc_sq[-1]==ir0_sq[-1]==ir1_sq[-1]==ir2_sq[-1] == diff_sq[-1]):
            join_var.append([int(len_sq[0]), float(gc_sq[0]), float(nc_sq[0]), float(ir0_sq[0]), float(ir1_sq[0]), float(ir2_sq[0]), float(diff_sq[0]) ]+nc_sq[1:])
        else:
            eprint([len_sq[-1],gc_sq[-1],nc_sq[-1],ir0_sq[-1],ir1_sq[-1],ir2_sq[-1] , diff_sq[-1]])
            eprint("ERROR in processing files, may lead to incorrect variable join")
            sys.exit() 

    return(join_var)

def join_average_variables(nc_ir0_avg_list,nc_ir1_avg_list,nc_ir2_avg_list):
    join_var=[]
    for ir0_sq, ir1_sq, ir2_sq in zip(nc_ir0_avg_list,nc_ir1_avg_list,nc_ir2_avg_list):
        if(ir0_sq[0]==ir1_sq[0]==ir2_sq[0]):
            join_var.append([ir0_sq[0],float(ir0_sq[1]), float(ir1_sq[1]), float(ir2_sq[1])])
        else:
            eprint([ir0_sq[0],float(ir0_sq[1]), float(ir1_sq[1]), float(ir2_sq[1])])
            eprint("ERROR in processing files, may lead to incorrect average variable join")
            sys.exit()
    return join_var

def regional_plots(joined_variables,save_folder):
    if not os.path.exists(save_folder):
        eprint("creating folder...")
        os.makedirs(save_folder)
    if not os.path.exists(save_folder+"_avg"):
        eprint("creating avg folder ...")
        os.makedirs(save_folder+"_avg")
    if not os.path.exists(save_folder+"_3d"):
        eprint("creating 3d folder ...")
        os.makedirs(save_folder+"_3d")
    if not os.path.exists(save_folder+"_boxplots"):
        eprint("creating boxplots folder...")
        os.makedirs(save_folder+"_boxplots")
    
    variables_boxplot=get_transformed_list(joined_variables)
    boxplot_genome(variables_boxplot,"NC","Sequence_Length","Program","Genome",save_folder+"_boxplots")
    variables_boxplot=[x[:-1] for x in variables_boxplot]
    joined_variables=[x[:-1] for x in joined_variables]
    boxplot(variables_boxplot,"NC","Sequence_Length","Program", ["Realm","Kingdom","Phylum","Class","Order","Family","Genus"],save_folder+"_boxplots")
    joined_variables = [x for x in joined_variables if x[0]<400000]
    joined_variables_scatter = random.sample(joined_variables,int(len(joined_variables)/10))
    plot_scatter(joined_variables_scatter,["Sequence_Length","GC_Content"] ,["NC","NC_IR2","diff_ir"],["Realm","Kingdom","Phylum","Class","Order"],save_folder,save_folder+"_avg")
    plot_3d_scatter(joined_variables_scatter,"Sequence_Length","GC_Content", ["NC","NC_IR2","diff_ir"],["Realm","Kingdom","Phylum","Class","Order"],save_folder+"_3d")

def get_transformed_list(joined_variables):
    reshaped_list=[]
    for x in joined_variables:
        reshaped_list.append(x[0:2] + [x[3], "IR 0"] + x[7:])
        reshaped_list.append(x[0:2] + [x[4], "IR 1"] + x[7:])
        reshaped_list.append(x[0:2] + [x[5], "IR 2"] + x[7:])
    return reshaped_list


def get_df_boxplot(join_var,index,string_name): 
    join_var=[x for x in join_var if string_name in x[index]]
    join_var=[x for x in join_var if "incertae_sedis" not in x[index+1]]
    order=list(set([x[index+1] for x in join_var]))
    return join_var,order



def get_df(join_var,index):
    joined_variables=[x for x in join_var if "incertae_sedis" not in x[index]]
    n_ulem=len(list(set([x[index] for x in joined_variables if "incertae_sedis" not in x[index]])))
    df = pd.DataFrame.from_records(joined_variables)
    df.columns =["Sequence_Length","GC_Content", "NC","NC_IR0","NC_IR1","NC_IR2","diff_ir","Super_Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus"]
    return df, n_ulem

def create_length_line(join_var,order, index):
    len_lst=[]
    for x in order:
        len_avg=[]
        for y in join_var:
            if x == y[index]:
                len_avg.append(y[0])
        len_lst.append(sum(len_avg)/len(len_avg))
    return len_lst
    

def boxplot_genome(join_var,yvar,yvar2,hue,tax_level, save_folder):
    print(f"{bcolors.OKBLUE}Creating Genome Boxplot...{bcolors.ENDC}")
    hue_order = ["IR 0", "IR 1","IR 2"]
    savefolder=save_folder+'/'+tax_level
    if not os.path.exists(savefolder):
        eprint("creating tax folder for", tax_level ,"...")
        os.makedirs(savefolder)
    order=["ssDNA","dsDNA","ssRNA","dsRNA","mixedDNA"]  
    len_list=create_length_line(join_var,order,-1)
    df = pd.DataFrame.from_records(join_var)
    df.columns =["Sequence_Length","GC_Content","NC","Program","Super_Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus","Genome"]
    
    
    g2 = sns.lineplot(x=order, y=len_list, marker="o", markers=True,sort=False, palette="Set3",sizes=(.05, 0.5))
    g2.set_ylabel("Sequence Length",fontsize='small')
    g2.xaxis.set_tick_params(labelsize='x-small')
    g2.set_xticklabels(tax_level)
    g2.set_xticklabels(g2.get_xticklabels(), rotation=30)
    g2.set(yscale="log") 
    ax2 = plt.twinx()
    ax2.set_xlabel(tax_level,fontsize='small')
    ax2.set_ylabel(tax_level,fontsize='small')
    g = sns.boxplot(x=tax_level, y=yvar, data=df,order=order,hue=hue,hue_order=hue_order, palette="Set3",ax=ax2, showfliers = False)
    plt.legend( bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='xx-small', borderaxespad=0.)
    width=0.8
    n_levels = len(df[hue].unique())
    each_width = width / n_levels

    offsets = np.linspace(0, width - each_width, n_levels)
    offsets -= offsets.mean()
    pos = [x+o for x in np.arange(len(order)) for o in offsets]

    counts = df.groupby([tax_level,hue])[yvar].size()
    counts = counts.reindex(pd.MultiIndex.from_product([order,hue_order]))
    medians = df.groupby([tax_level,hue])[yvar].median()
    medians = medians.reindex(pd.MultiIndex.from_product([order,hue_order]))

    for p,n,m in zip(pos,counts,medians):
        if not np.isnan(m):
            g.annotate('{:.0f}'.format(n), xy=(p, m), xycoords='data', ha='center', va='bottom',fontsize=4)
    
    save_string=tax_level+"_"+yvar+"_vs_"+yvar2+".pdf"
    plt.savefig(savefolder+"/"+save_string,  bbox_inches='tight')
    plt.clf()
    print(f"{bcolors.OKBLUE}Genome boxplot successfully stored in:{bcolors.ENDC}",savefolder+"/"+save_string)

def create_length_line_2(join_variables,order):
    join_variables=[x for x in join_variables if x[3]=="IR 0"]
    len_lst=[]
    for x in order:
        for y in join_variables[0]:
            if x == y:
                index = join_variables[0].index(x)
                break
    ordr=[]
    for x in order:
        len_avg=[]
        for y in join_variables:
            if x == y[index]:
                len_avg.append(y[0])
        ordr.append(x)
        len_lst.append(sum(len_avg)/len(len_avg))
    return ordr, len_lst

def boxplot(join_var,yvar,yvar2,hue,tax_level, save_folder):
    print(f"{bcolors.OKBLUE}Creating all Taxonomic Boxplots...{bcolors.ENDC}")
    hue_order = ["IR 0", "IR 1","IR 2"]
    for idx,tx_level in zip(range(4,11),tax_level): 
        savefolder=save_folder+'/'+tx_level
        if not os.path.exists(savefolder):
            eprint("creating tax folder for", tx_level ,"...")
            os.makedirs(savefolder)
        tax_enum_list=list(set([x[idx] for x in join_var]))
        for tax in tax_enum_list:
            join_variables, order = get_df_boxplot(join_var,idx,tax)
            if join_variables:
                df = pd.DataFrame.from_records(join_variables)
                df.columns =["Sequence_Length","GC_Content","NC","Program","Super_Realm","Realm","Kingdom","Phylum","Class","Order","Family","Genus"]
                order, len_list=create_length_line_2(join_variables,order)
                g2 = sns.lineplot(x=order, y=len_list, marker="o", markers=True,sort=False, palette="Set3",sizes=(.05, 0.5))               
                g2.set_ylabel("Sequence Length",fontsize='small')
                g2.xaxis.set_tick_params(labelsize='x-small')
                g2.set_xticklabels(g2.get_xticklabels(), rotation=30)

                g2.set(yscale="log")            
                ax2 = plt.twinx()
                ax2.set_xlabel(tx_level,fontsize='small')
                ax2.set_ylabel(tx_level,fontsize='small')
                g = sns.boxplot(x=tx_level, y=yvar, data=df,order=order,hue=hue,hue_order=hue_order,ax=ax2, palette="Set3", showfliers = False)
                plt.legend( bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='xx-small', borderaxespad=0.)

                ########################################
                width=0.8
                n_levels = len(df[hue].unique())
                each_width = width / n_levels

                offsets = np.linspace(0, width - each_width, n_levels)
                offsets -= offsets.mean()
                pos = [x+o for x in np.arange(len(order)) for o in offsets]

                counts = df.groupby([tx_level,hue])[yvar].size()
                counts = counts.reindex(pd.MultiIndex.from_product([order,hue_order]))
                medians = df.groupby([tx_level,hue])[yvar].median()
                medians = medians.reindex(pd.MultiIndex.from_product([order,hue_order]))

                for p,n,m in zip(pos,counts,medians):
                    if not np.isnan(m):
                        g.annotate('{:.0f}'.format(n), xy=(p, m), xycoords='data', ha='center', va='bottom',fontsize=4)
                
                save_string=tax+"_"+yvar+"_vs_"+yvar2+".pdf"
                plt.savefig(savefolder+"/"+save_string,  bbox_inches='tight')
                plt.clf()
                plt.close(1)
        print(f"{bcolors.OKBLUE} All taxonomic boxplots of the {bcolors.ENDC}",tx_level,f"{bcolors.OKBLUE} were successfully stored in:{bcolors.ENDC}",savefolder+"/")

            ########################################

def plot_scatter(join_var,list_x,list_y, group_by, save_folder,save_folder2 ):
    print(f"{bcolors.OKBLUE}Creating all 2d scatter plots...{bcolors.ENDC}")
    for metric_axis_x in list_x:
        for metric_axis_y in list_y:
            cnt=8
            for group_str in group_by:
                df,n_col = get_df(join_var,cnt)
                colors=sns.color_palette("husl", n_col)
                palette=sns.color_palette(colors)
                g = sns.scatterplot(x=metric_axis_x, y=metric_axis_y, hue=group_str , data=df, palette=palette)
                g.set(xscale="log")
                g.set_xlabel(metric_axis_x.replace('_', ' ').title(),fontsize='medium')
                g.set_ylabel(metric_axis_y.replace('_', ' ').title(),fontsize='medium')
                save_string=metric_axis_x+"_vs_"+metric_axis_y+"_"+group_str+".pdf"
                plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='xx-small', borderaxespad=0.)
                plt.xticks(fontsize='medium')
                plt.yticks(fontsize='medium')
                plt.savefig(save_folder+"/"+save_string,  bbox_inches='tight')
                plt.clf()

                gb=df.groupby(group_str)
                mean = gb.mean()
                
                
                groups = [name for name,unused_df in gb]
                sdevs = df.groupby(group_str).std()
                g = sns.scatterplot(x=metric_axis_x, y=metric_axis_y, marker='o', s=50, fc='none', ec=palette,label=groups, data=mean)
                for (_, mean), (_, sdev), color in zip(mean.iterrows(), sdevs.iterrows(), palette):
                    for sdev_mult in [1, 2, 3]:
                        ellipse = Ellipse((mean[metric_axis_x], mean[metric_axis_y]), width=sdev[metric_axis_x] * sdev_mult,
                                        height=sdev[metric_axis_y] * sdev_mult,facecolor=color, alpha=0.2 if sdev_mult == 1 else 0.1)
                        g.add_patch(ellipse)
                plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='xx-small', borderaxespad=0.)
                plt.xticks(fontsize='small')
                plt.yticks(fontsize='small')
                
                plt.savefig(save_folder2+"/"+save_string)
                plt.clf()
                plt.close(1)
                cnt+=1
    print(f"{bcolors.OKBLUE} All 2d scatter plots were successfully stored in:{bcolors.ENDC}",save_folder+"/")

def plot_3d_scatter(join_var,metric_axis_x,metric_axis_y, list_z, group_by, save_folder):
    print(f"{bcolors.OKBLUE}Creating all 3d scatter plots...{bcolors.ENDC}")
    for metric_axis_z in list_z:
        cnt=8
        for group_str in group_by:
            ax = plt.axes(projection ="3d")
            save_string=metric_axis_x+"_vs_"+metric_axis_y+"_vs_"+metric_axis_z+"_"+group_str+".pdf"
            df,nc=get_df(join_var,cnt)
            colors=sns.color_palette("husl", nc)
            n=0
            for s in df[group_str].unique():
                ax.scatter(df[metric_axis_y][df[group_str]==s],df[metric_axis_x][df[group_str]==s],df[metric_axis_z][df[group_str]==s],label=s, color=colors[n])
                n+=1
            ax.invert_yaxis()

            ax.view_init(35, 45)
            ax.set_xlabel('GC Content',fontsize='medium')
            ax.set_ylabel('Sequence Length',fontsize='medium')
            ax.set_zlabel(metric_axis_z,fontsize='medium')
            
            plt.xticks(fontsize='small')
            plt.yticks(fontsize='small')
            
            plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize='xx-small', borderaxespad=0.)
            plt.savefig(save_folder+"/"+save_string,  bbox_inches='tight')
            plt.clf()
            plt.close(1)
            cnt+=1
    print(f"{bcolors.OKBLUE} All 3d scatter plots were successfully stored in:{bcolors.ENDC}",save_folder+"/")
    


def max_min(list_nc):
    nc_v=[x[2] for x in list_nc]
    std = statistics.stdev(nc_v)
    average = sum(nc_v)/len(nc_v)
    max_v=average+(3*std)
    min_v=average-(3*std)

    return max_v, min_v

def filter_outliers(list_len, list_gc, list_nc, list_ir0, list_ir1, list_ir2):
    max_nc, min_nc = max_min(list_nc)
    max_ir0, min_ir0 = max_min(list_ir0)
    max_ir1, min_ir1 = max_min(list_ir1)
    max_ir2, min_ir2 = max_min(list_ir2)
    new_len_list=[]
    new_gc_list=[]
    new_nc_list=[]
    new_ir0_list=[]
    new_ir1_list=[]
    new_ir2_list=[]
    for ln, gc, nc, ir0, ir1, ir2 in zip(list_len, list_gc, list_nc, list_ir0, list_ir1, list_ir2):
        
        if (ln[1]==gc[1]==nc[1]==ir0[1] == ir1[1] == ir2[1]) and (nc[2]>min_nc and nc[2]<max_nc) and (ir0[2]>min_ir0 and ir0[2]<max_ir0) and (ir1[2]>min_ir1 and ir1[2]<max_ir1) and (ir2[2]>min_ir2 and ir2[2]<max_ir2):
            new_len_list.append(ln)
            new_gc_list.append(gc)
            new_nc_list.append(nc)
            new_ir0_list.append(ir0)
            new_ir1_list.append(ir1)
            new_ir2_list.append(ir2)
        elif(ln[1]==gc[1]==nc[1]==ir0[1] == ir1[1] == ir2[1])==False:
            eprint("ERROR files do not coincide")
            eprint(ln, gc, nc, ir0, ir1, ir2)
            sys.exit()    

    return  new_len_list, new_gc_list, new_nc_list, new_ir0_list, new_ir1_list, new_ir2_list

def hasNumbers(inputString):
    return any(char.isdigit() for char in inputString)



def join_gen(meta_info, join_var):
    gen_m=[[x[2], x[10]] for x in meta_info]
    gen_m = [list(x) for x in set(tuple(x) for x in gen_m)]
    rf = read_file("../aux/fix_genome.txt")
    rm=[x[0] for x in rf]
    gen_m = [x for x in gen_m if x[-1] not in rm]
    [gen_m.append([x[1],x[0]]) for x in rf]    
    gen_m = [list(x) for x in set(tuple(x) for x in gen_m)]
    gen_m = [["ssRNA",x[1]]   if x[0]=="sscRNA" else x for x in gen_m ]
    jn_var=[]
    for g in gen_m: 
        for x in join_var:
            if g[-1]==x[-1]:
                jn_var.append(x + [g[0]])
    return jn_var

def to_dict(lst):
    d = {}
    for el,i in zip(lst, range(0,len(lst))):
        d[el] = i
    return d

def create_distribution(nc_taxa):
    
    print("Creating distribution of genus plot...")
    l=[x[2] for x in nc_taxa]
    occurance=[[x,l.count(x)] for x in set(l)]
    labels=[x[0] for x in occurance]
    values=[x[1] for x in occurance]
    values.sort(reverse=True) 
    max_value=values[0]
    min_value=values[30]

    
    lObjectsALLcnts = [max_value/3 if x[1]==max_value else x[1] for x in occurance]
    lObjectsALLlbls = [x[0] + "-" + str(x[1]) if x[1]>min_value else "" for x in occurance ]

    iN = len(lObjectsALLcnts)
    arrCnts = np.array(lObjectsALLcnts)

    theta=np.arange(0,2*np.pi,2*np.pi/iN)
    width = (2*np.pi)/iN *0.9

    fig = plt.figure(figsize=(8, 8))
    ax = fig.add_axes([0.1, 0.1, 0.75, 0.75], polar=True)
    bars = ax.bar(theta ,arrCnts, width=width,color='g', bottom=50)
    ax.set_xticks(theta)
    plt.axis('off')
    rotations = [np.degrees(i) for i in theta]
    for i in rotations: i = int(i)
    for x, bar, rotation, label in zip(theta, bars, rotations, lObjectsALLlbls):
        height = bar.get_height() + 50
        ax.text(x + bar.get_width(), height, label, fontsize=3, ha='center', va='bottom', rotation=rotation)
    plt.savefig("../plots/occurances.pdf", dpi=5000)
    plt.clf()
    plt.close(1)
    print("The program successfully stored the distribution of genus plot! \n To view the plot open occurances.pdf in the /plots folder.")


if __name__ == "__main__": 
    ## Paths
    #Save Lists Path
    xls_path="../xlslist"
    print(f"{bcolors.OKGREEN}Starting IR Analysis...{bcolors.ENDC}")

    #Metagenomic Information Paths
    organism_path="../VirusDB/ViralSeq_Org.info"
    genomic_type_path="../VirusDB/ViralSeq_Genome.info"
    
    #Features Paths
    NC_FILE="../reports/REPORT_COMPLEXITY_NC_OTHER_3"
    len_seq_path="../reports/REPORT_SEQ_LEN"
    GC_path="../reports/REPORT_SEQ_GC"
    IR_0="../reports/Report_NC_IR_OPTIMAL_0"
    IR_1="../reports/Report_NC_IR_OPTIMAL_1"
    IR_2="../reports/Report_NC_IR_OPTIMAL_2"
    

    

    ##Process meta information
    print(f"{bcolors.OKGREEN}Processing genomic meta information...{bcolors.ENDC}")
    rg_list = read_genomic_type_path(genomic_type_path)
    ro_list = read_organism_path(organism_path)
    ll = match_files(rg_list,ro_list)
    meta_info = process_meta_information(ll)
    meta_info = read_and_correct("../VirusDB/correct_taxa.txt", meta_info)
    tax_ign=read_to_list("../VirusDB/ignoretax.txt")
    meta_info=filter_from_ll(meta_info, tax_ign)
    organism_names = [x[3:-1] for x in meta_info]
    organism_names=filter_organism_list(organism_names)    
    meta_info=filter_meta_info(meta_info)



    new_tree=[]
    unique_list = process_organism_lists(organism_names)
    eprint("org:",len(organism_names))
    eprint("meta_info:",len(meta_info))
    for x,y in zip(organism_names, meta_info):
        if x:
            new_tree.append(y + [x[7]])
        else:
            new_tree.append(y +[""])
    

    ##Process Features
    print(f"{bcolors.OKGREEN}Processing Features...{bcolors.ENDC}")
    len_seq=read_file(len_seq_path) # Sequence Length
    gc_seq=read_file(GC_path) # GC-Content
    nc_file=nc_process_no_floor(NC_FILE) #NC from the best Compression Level
    best_nc_for_file_0 = nc_process_no_floor(IR_0) #Context model Compression
    best_nc_for_file_1 = nc_process_no_floor(IR_1) #Mixed Compression
    best_nc_for_file_2 = nc_process_no_floor(IR_2) #IR-based Compression

    # Filter outliers
    print(f"{bcolors.OKGREEN}Filtering Outliers...{bcolors.ENDC}")
    len_seq, gc_seq, nc_file, best_nc_for_file_0, best_nc_for_file_1, best_nc_for_file_2 = filter_outliers(len_seq, gc_seq, nc_file,best_nc_for_file_0, best_nc_for_file_1,best_nc_for_file_2)
    eprint("----------------------")
    eprint("len_seq: ",len(len_seq))
    eprint("gc_seq: ",len(gc_seq))
    eprint("nc_file: ",len(nc_file))
    eprint("best_nc_for_file_0: ",len(best_nc_for_file_0))
    eprint("best_nc_for_file_1: ",len(best_nc_for_file_1))
    eprint("best_nc_for_file_2: ",len(best_nc_for_file_2))
    
    #Diff between NC(IR0)-NC(IR1)
    diff_nc = nc_diff(best_nc_for_file_0, best_nc_for_file_1)

    # Performing Further Pre-processing of Features 
    # Removing Features of Incomplete Taxonomic Metadata.
    print(f"{bcolors.OKGREEN}Removing Features with Incomplete Taxonomic Metadata...{bcolors.ENDC}")
    len_taxa,_ = join_nc_taxa(len_seq, new_tree)
    gc_taxa,_ = join_nc_taxa(gc_seq, new_tree)
    nc_taxa,_ = join_nc_taxa(nc_file, new_tree)
    nc_taxa_ir1,_ = join_nc_taxa(best_nc_for_file_1, new_tree)
    nc_taxa_ir2,_ = join_nc_taxa(best_nc_for_file_2, new_tree)
    nc_taxa_ir0,_ = join_nc_taxa(best_nc_for_file_0, new_tree)
    diff_nc_taxa,_ = join_nc_taxa(diff_nc, new_tree)

    #Creationg of the occurance of genus plot
    create_distribution(nc_taxa)
    
    eprint("----------------------")
    eprint("len_taxa: ",len(len_taxa))
    eprint("gc_taxa: ",len(gc_taxa))
    eprint("nc_taxa: ",len(nc_taxa))
    eprint("nc_taxa_ir0: ",len(nc_taxa_ir0))
    eprint("nc_taxa_ir1: ",len(nc_taxa_ir1))
    eprint("nc_taxa_ir2: ",len(nc_taxa_ir2))
    eprint("diff_nc_taxa: ",len(diff_nc_taxa))
    eprint("----------------------")
   
    ## Creating Top Table Lists 
    print(f"{bcolors.OKGREEN}Creating Top Table Lists...{bcolors.ENDC}")
    len_taxa_process = [[a[1]] + a[3:-1] for a in len_taxa]
    gc_taxa_process = [[a[1]] + a[3:-1] for a in gc_taxa]
    nc_taxa_process = [[a[1]] + a[3:-1] for a in nc_taxa]
    nc_taxa_ir0_process = [[a[1]] + a[3:-1] for a in nc_taxa_ir0 ]
    nc_taxa_ir1_process = [[a[1]] + a[3:-1] for a in nc_taxa_ir1 ]
    nc_taxa_ir2_process = [[a[1]] + a[3:-1] for a in nc_taxa_ir2 ]
    diff_nc_taxa_process = [[a[1]] + a[3:-1] for a in diff_nc_taxa ]
    
    nc_taxa_fused_list = list_fusion(nc_taxa_process, len_taxa_process, gc_taxa_process)
    nc_taxa_fused_list = join_gen(meta_info,nc_taxa_fused_list)
    
    nc_ir2_taxa_fused_list = list_fusion(nc_taxa_ir2_process, len_taxa_process, gc_taxa_process)
    nc_ir2_taxa_fused_list = [ [1-a[0]]+a[1:] for a in nc_ir2_taxa_fused_list if (1-a[0])>0 ]
    nc_ir2_taxa_fused_list = join_gen(meta_info,nc_ir2_taxa_fused_list)
    
    diff_nc_taxa_taxa_fused_list = list_fusion(diff_nc_taxa_process, len_taxa_process, gc_taxa_process)
    diff_nc_taxa_taxa_fused_list=[ a for a in diff_nc_taxa_taxa_fused_list if a[0]>0]
    diff_nc_taxa_taxa_fused_list = join_gen(meta_info,diff_nc_taxa_taxa_fused_list)

    eprint("NC fused list len :", len(nc_taxa_fused_list))
    eprint("IR2 fused list len :", len(nc_ir2_taxa_fused_list))
    eprint("DIFF fused list len :", len(diff_nc_taxa_taxa_fused_list))
    
    create_features_list(nc_taxa_fused_list, xls_path, "nc_len_gc")
    create_features_list(nc_ir2_taxa_fused_list, xls_path, "ir2_len_gc")
    create_features_list(diff_nc_taxa_taxa_fused_list, xls_path, "diff_len_gc")

    nc_taxa_list = [[a[1]] + a[3:-1] for a in nc_taxa]
    nc_taxa_ir2_list = [[1-a[1]] + a[3:-1] for a in nc_taxa_ir2 if (1-a[1])>0]
    diff_nc_taxa_list = [[a[1]] + a[3:-1] for a in diff_nc_taxa if a[1]>0]

    create_lists(nc_taxa_list,xls_path, "nc")
    create_lists(nc_taxa_ir2_list,xls_path, "ir2")
    create_lists(diff_nc_taxa_list,xls_path, "diff_ir")
    
    ##Regional Plots
    print(f"{bcolors.OKGREEN}Performing Regional Plots...{bcolors.ENDC}")
    join_var = join_variables(len_taxa_process, gc_taxa_process, nc_taxa_process,nc_taxa_ir0_process, nc_taxa_ir1_process, nc_taxa_ir2_process, diff_nc_taxa_process)
    eprint("len join_var: ", len(join_var))
    join_var = join_gen(meta_info,join_var)
    eprint("len join_var: ", len(join_var))
    regional_plots(join_var, "../plots/agregation_plots")
    print(f"{bcolors.OKGREEN}Successfully completed IR Analysis!{bcolors.ENDC}")

    ##Phylogenetic Trees Leafs Ending in Different Taxonomic Groups.
    # print(f"{bcolors.OKGREEN}Creating All Phylogenetic Trees...{bcolors.ENDC}")
    # nc_avg_list = tree_average(nc_taxa_process)
    # nc_ir2_avg_list = tree_average(nc_taxa_ir2_list)
    # diff_nc_avg_list = tree_average(diff_nc_taxa_list)
    # various_trees(organism_names, nc_avg_list, "Viruses", "nc_tree", False,"NC","../plots/phylo_trees")
    # various_trees(organism_names, nc_ir2_avg_list, "Viruses", "nc_ir2_tree", False,"IR0","../plots/phylo_trees") #maybe invert scale
    # various_trees(organism_names, diff_nc_avg_list, "Viruses", "diff_nc_tree",False,"DIFF","../plots/phylo_trees")
    # ##Website Group-specific Phylogenetic Trees
    # various_trees_2(organism_names, nc_avg_list,"../plots/trees_2")
    # print(f"{bcolors.OKGREEN}IR Analysis Successfully Complete...{bcolors.ENDC}")

    
 
    
